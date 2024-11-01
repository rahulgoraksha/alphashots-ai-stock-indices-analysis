import yfinance as yf
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import seaborn as sns
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import pytz

class CurrencyAnalyzer:
    def __init__(self, symbol="EURINR=X", start_date="2023-01-01", end_date="2024-09-30"):
        self.symbol = symbol
        self.start_date = start_date
        self.end_date = end_date
        self.data = None
        self.daily_data = None
        self.weekly_data = None
        
    def fetch_data(self):
        try:
            df = yf.download(self.symbol, start=self.start_date, end=self.end_date)
            
            if df.empty:
                raise ValueError("No data retrieved from Yahoo Finance")
            
            print("Available columns:", df.columns.tolist())
            
            # Remove multi-index if present
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)
            
            # Rename columns if they're different
            if 'Adj Close' in df.columns:
                df['Close'] = df['Adj Close']
            
            # Handle missing Volume column
            if 'Volume' not in df.columns:
                df['Volume'] = 0
                print("Warning: Volume data not available, using placeholder values")
            
            # Validate data completeness
            missing_dates = pd.date_range(start=self.start_date, end=self.end_date).difference(df.index)
            if len(missing_dates) > 0:
                print(f"Warning: Missing data for {len(missing_dates)} dates")
                
            self.data = df
            print(f"Successfully downloaded {len(df)} records")
            return True
            
        except Exception as e:
            print(f"Error fetching data: {e}")
            return False

    def prepare_timeframes(self):
        try:
            # Daily data
            self.daily_data = self.calculate_technical_indicators(self.data)
            
            # Weekly data - proper OHLC aggregation
            available_columns = {}
            for col in ['Open', 'High', 'Low', 'Close', 'Volume']:
                if col in self.data.columns:
                    available_columns[col] = 'first' if col == 'Open' else 'last' if col == 'Close' else 'max' if col == 'High' else 'min' if col == 'Low' else 'sum'
            
            if not available_columns:
                raise ValueError("No valid columns found for aggregation")
            
            self.weekly_data = self.data.resample('W').agg(available_columns).dropna()
            self.weekly_data = self.calculate_technical_indicators(self.weekly_data)
            
        except Exception as e:
            print(f"Error preparing timeframes: {e}")
            raise

    def calculate_technical_indicators(self, df):
        df = df.copy()
        
        # Verify we have the required columns
        required_columns = ['Close']
        if not all(col in df.columns for col in required_columns):
            raise ValueError(f"Missing required columns. Found columns: {df.columns.tolist()}")
        
        # Simple Moving Average (20-day)
        df['MA20'] = df['Close'].rolling(window=20, min_periods=20).mean()
        
        # Bollinger Bands
        rolling_mean = df['Close'].rolling(window=20, min_periods=20).mean()
        rolling_std = df['Close'].rolling(window=20, min_periods=20).std()
        df['BB_middle'] = rolling_mean
        df['BB_upper'] = rolling_mean + (2 * rolling_std)
        df['BB_lower'] = rolling_mean - (2 * rolling_std)
        
        # Enhanced CCI calculation
        if all(col in df.columns for col in ['High', 'Low']):
            typical_price = (df['High'] + df['Low'] + df['Close']) / 3
        else:
            print("Warning: Using Close price only for CCI calculation")
            typical_price = df['Close']
            
        rolling_mean_tp = typical_price.rolling(window=20, min_periods=20).mean()
        mean_deviation = abs(typical_price - rolling_mean_tp).rolling(window=20, min_periods=20).mean()
        df['CCI'] = np.where(
            mean_deviation != 0,
            (typical_price - rolling_mean_tp) / (0.015 * mean_deviation),
            np.nan
        )
        
        # Add validation metrics
        df['Valid_Signals'] = df[['MA20', 'BB_middle', 'CCI']].notna().all(axis=1)
        
        return df

    def generate_signals(self, df, target_date):
        signals = {}
        confidence = {}
        
        try:
            # Get target date data and previous day for trend
            target_data = df.loc[target_date]
            prev_date = df.index[df.index.get_loc(target_date) - 1]
            prev_data = df.loc[prev_date]
            
            # Moving Average Signal with trend strength
            ma_trend = (target_data['Close'] - target_data['MA20']) / target_data['MA20'] * 100
            if target_data['Close'] > target_data['MA20']:
                signals['MA'] = 'BUY'
                confidence['MA'] = min(abs(ma_trend) / 2, 100)
            elif target_data['Close'] < target_data['MA20']:
                signals['MA'] = 'SELL'
                confidence['MA'] = min(abs(ma_trend) / 2, 100)
            else:
                signals['MA'] = 'NEUTRAL'
                confidence['MA'] = 0
            
            # Bollinger Bands Signal with distance-based confidence
            bb_position = (target_data['Close'] - target_data['BB_middle']) / (target_data['BB_upper'] - target_data['BB_middle'])
            if target_data['Close'] > target_data['BB_upper']:
                signals['BB'] = 'SELL'
                confidence['BB'] = min((bb_position - 1) * 100, 100)
            elif target_data['Close'] < target_data['BB_lower']:
                signals['BB'] = 'BUY'
                confidence['BB'] = min(abs(bb_position) * 100, 100)
            else:
                signals['BB'] = 'NEUTRAL'
                confidence['BB'] = 0
            
            # CCI Signal with strength consideration
            if target_data['CCI'] > 100:
                signals['CCI'] = 'SELL'
                confidence['CCI'] = min((target_data['CCI'] - 100) / 2, 100)
            elif target_data['CCI'] < -100:
                signals['CCI'] = 'BUY'
                confidence['CCI'] = min(abs(target_data['CCI'] + 100) / 2, 100)
            else:
                signals['CCI'] = 'NEUTRAL'
                confidence['CCI'] = 0
                
        except KeyError:
            print(f"Warning: Data not available for {target_date}")
            signals = {'MA': 'N/A', 'BB': 'N/A', 'CCI': 'N/A'}
            confidence = {'MA': 0, 'BB': 0, 'CCI': 0}
        
        return signals, confidence

    def create_analysis_plots(self, df, timeframe, output_dir='.'):
        plt.style.use('classic')
        fig, axes = plt.subplots(3, 1, figsize=(15, 15))
        
        colors = {
            'price': '#2E86C1',
            'ma': '#F39C12',
            'bb': ['#E74C3C', '#2ECC71', '#3498DB'],
            'cci': '#8E44AD'
        }
        
        # Plot 1: Price and MA
        axes[0].plot(df.index, df['Close'], label='Close Price', color=colors['price'])
        axes[0].plot(df.index, df['MA20'], label='20-day MA', color=colors['ma'])
        axes[0].set_title(f'EUR/INR Price and Moving Average ({timeframe})')
        axes[0].legend()
        axes[0].grid(True, alpha=0.3)
        
        # Plot 2: Bollinger Bands
        axes[1].plot(df.index, df['Close'], label='Close Price', color=colors['price'])
        axes[1].plot(df.index, df['BB_upper'], label='Upper BB', color=colors['bb'][0], linestyle='--')
        axes[1].plot(df.index, df['BB_middle'], label='Middle BB', color=colors['bb'][2])
        axes[1].plot(df.index, df['BB_lower'], label='Lower BB', color=colors['bb'][1], linestyle='--')
        axes[1].fill_between(df.index, df['BB_upper'], df['BB_lower'], alpha=0.1, color=colors['bb'][2])
        axes[1].set_title(f'Bollinger Bands ({timeframe})')
        axes[1].legend()
        axes[1].grid(True, alpha=0.3)
        
        # Plot 3: CCI
        axes[2].plot(df.index, df['CCI'], label='CCI', color=colors['cci'])
        axes[2].axhline(y=100, color='r', linestyle='--', alpha=0.5)
        axes[2].axhline(y=-100, color='g', linestyle='--', alpha=0.5)
        axes[2].axhline(y=0, color='gray', linestyle='-', alpha=0.3)
        axes[2].fill_between(df.index, 100, df['CCI'], where=(df['CCI'] > 100), color='r', alpha=0.1)
        axes[2].fill_between(df.index, -100, df['CCI'], where=(df['CCI'] < -100), color='g', alpha=0.1)
        axes[2].set_title(f'Commodity Channel Index ({timeframe})')
        axes[2].legend()
        axes[2].grid(True, alpha=0.3)
        
        plt.tight_layout()
        filename = f'{output_dir}/technical_analysis_{timeframe}.png'
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        plt.close()
        return filename

    def save_results(self, daily_signals, daily_confidence, weekly_signals, weekly_confidence, filename):
        results = pd.DataFrame({
            'Timeframe': ['Daily', 'Weekly'],
            'MA Signal': [daily_signals['MA'], weekly_signals['MA']],
            'MA Confidence': [f"{daily_confidence['MA']:.1f}%", f"{weekly_confidence['MA']:.1f}%"],
            'BB Signal': [daily_signals['BB'], weekly_signals['BB']],
            'BB Confidence': [f"{daily_confidence['BB']:.1f}%", f"{weekly_confidence['BB']:.1f}%"],
            'CCI Signal': [daily_signals['CCI'], weekly_signals['CCI']],
            'CCI Confidence': [f"{daily_confidence['CCI']:.1f}%", f"{weekly_confidence['CCI']:.1f}%"]
        })
        
        # Export results to Excel
        results.to_csv(filename, index=False)
        print(f"\nResults saved to {filename}")
        print("\nSignals Summary:")
        print(results.to_string())

    def save_to_excel(self, daily_signals, daily_confidence, weekly_signals, weekly_confidence, filename):
        excel_filename = f"Currency_Analysis.xlsx"
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            self._create_summary_sheet(writer, daily_signals, daily_confidence, weekly_signals, weekly_confidence)
            self._create_data_sheet(writer, self.daily_data, 'Daily_Data')
            self._create_data_sheet(writer, self.weekly_data, 'Weekly_Data')
            self._create_technical_analysis_sheet(writer)
            self._create_raw_data_sheet(writer)
            
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        print(f"\nExcel file saved as: {excel_filename}")
        return excel_filename

    def _create_summary_sheet(self, writer, daily_signals, daily_confidence, weekly_signals, weekly_confidence):
        summary_data = {
            'Timeframe': ['Daily', 'Weekly'],
            'MA Signal': [daily_signals.get('MA', 'N/A'), weekly_signals.get('MA', 'N/A')],
            'MA Confidence': [f"{daily_confidence.get('MA', 0):.1f}%", f"{weekly_confidence.get('MA', 0):.1f}%"],
            'BB Signal': [daily_signals.get('BB', 'N/A'), weekly_signals.get('BB', 'N/A')],
            'BB Confidence': [f"{daily_confidence.get('BB', 0):.1f}%", f"{weekly_confidence.get('BB', 0):.1f}%"],
            'CCI Signal': [daily_signals.get('CCI', 'N/A'), weekly_signals.get('CCI', 'N/A')],
            'CCI Confidence': [f"{daily_confidence.get('CCI', 0):.1f}%", f"{weekly_confidence.get('CCI', 0):.1f}%"]
        }
        summary_df = pd.DataFrame(summary_data)
        
        metadata_df = pd.DataFrame({
            'Parameter': ['Currency Pair', 'Start Date', 'End Date', 'Analysis Date'],
            'Value': [self.symbol, self.start_date, self.end_date, datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        })

        metadata_df.to_excel(writer, sheet_name='Summary', startrow=0, index=False)
        summary_df.to_excel(writer, sheet_name='Summary', startrow=len(metadata_df) + 2, index=False)

    def _create_data_sheet(self, writer, df, sheet_name):
        if df is None or df.empty:
            print(f"Warning: No data available for {sheet_name}")
            return

        df_excel = df.reset_index()
        numeric_columns = df_excel.select_dtypes(include=[np.number]).columns
        df_excel[numeric_columns] = df_excel[numeric_columns].round(4)
        df_excel.to_excel(writer, sheet_name=sheet_name, index=False)

    def _create_technical_analysis_sheet(self, writer):
        analysis_data = {
            'Indicator': ['Moving Average (MA20)', 'Bollinger Bands', 'CCI (Commodity Channel Index)'],
            'Description': [
                'A 20-day moving average that helps identify trends. Values above MA suggest bullish trend, below suggest bearish trend.',
                'Consists of a middle band (20-day MA) and upper/lower bands (2 standard deviations). Helps identify overbought/oversold conditions.',
                'Measures current price level relative to average price level over time. Values above +100 suggest overbought, below -100 suggest oversold.'
            ],
            'Signal Generation': [
                'BUY: Price > MA20\nSELL: Price < MA20\nNEUTRAL: Price ≈ MA20',
                'BUY: Price < Lower Band\nSELL: Price > Upper Band\nNEUTRAL: Price between bands',
                'BUY: CCI < -100\nSELL: CCI > +100\nNEUTRAL: -100 < CCI < +100'
            ]
        }
        
        analysis_df = pd.DataFrame(analysis_data)
        analysis_df.to_excel(writer, sheet_name='Technical_Analysis', index=False)

    def _create_raw_data_sheet(self, writer):
        if self.data is None or self.data.empty:
            print("Warning: No raw data available")
            return

        # Write raw data to Excel
        self.data.to_excel(writer, sheet_name='Raw_Data', index=True)
        
    def calculate_indicators(self, date, timeframe='1d'):
        """
        Calculate technical indicators for a specific date and timeframe.
        """
        if timeframe == '1d':
            df = self.daily_data
        elif timeframe == '1w':
            df = self.weekly_data
        else:
            raise ValueError("Invalid timeframe. Use '1d' for daily or '1w' for weekly.")

        # Ensure the date exists in the dataframe
        if date not in df.index:
            closest_date = df.index[df.index <= date][-1]
            print(f"Warning: Data not available for {date}. Using closest available date: {closest_date}")
            date = closest_date

        # Calculate indicators
        ma20 = df.loc[date, 'MA20']
        bb_upper = df.loc[date, 'BB_upper']
        bb_middle = df.loc[date, 'BB_middle']
        bb_lower = df.loc[date, 'BB_lower']
        cci = df.loc[date, 'CCI']

        return {
            'Date': date,
            'Close': df.loc[date, 'Close'],
            'MA20': ma20,
            'BB_upper': bb_upper,
            'BB_middle': bb_middle,
            'BB_lower': bb_lower,
            'CCI': cci
        }

def main():
    # Initialize analyzer
    analyzer = CurrencyAnalyzer()
    
    # Fetch and prepare data
    if not analyzer.fetch_data():
        print("Failed to fetch data. Exiting.")
        return
    
    try:
        analyzer.prepare_timeframes()
    except Exception as e:
        print(f"Error in prepare_timeframes: {e}")
        return
    
    # Target date
    target_date = datetime.strptime("2024-09-30", "%Y-%m-%d")
    
    # Calculate indicators for one day from Sept 30, 2024
    one_day_later = pytz.utc.localize(target_date + timedelta(days=1))
    daily_indicators = analyzer.calculate_indicators(one_day_later, '1d')
    
    # Calculate indicators for one week from Sept 30, 2024
    one_week_later = pytz.utc.localize(target_date + timedelta(weeks=1))
    weekly_indicators = analyzer.calculate_indicators(one_week_later, '1w')
    
    # Print results
    print("\nTechnical Analysis Results:")
    print("\nOne day from Sept 30, 2024:")
    for key, value in daily_indicators.items():
        print(f"{key}: {value}")
    
    # Find the last available date if target date is not available
    if target_date not in analyzer.daily_data.index:
        target_date = analyzer.daily_data.index[-1]
        print(f"Warning: Using last available date: {target_date}")
    
    daily_signals, daily_confidence = analyzer.generate_signals(analyzer.daily_data, target_date)
    
    # For weekly data, find the last week that contains our target date
    weekly_target_date = analyzer.weekly_data.index[analyzer.weekly_data.index <= target_date][-1]
    weekly_signals, weekly_confidence = analyzer.generate_signals(analyzer.weekly_data, weekly_target_date)
    
    # Create plots
    analyzer.create_analysis_plots(analyzer.daily_data.tail(100), '1d')
    analyzer.create_analysis_plots(analyzer.weekly_data.tail(52), '1w')
    
    # Save results
    results = analyzer.save_results(
        daily_signals,
        daily_confidence,
        weekly_signals,
        weekly_confidence,
        f"Currency_Analysis.csv"
    )
    
    # Save to Excel
    try:
        excel_filename = analyzer.save_to_excel(
            daily_signals,
            daily_confidence,
            weekly_signals,
            weekly_confidence,
            f"Currency_Analysis.xlsx"
        )
        print(f"Excel file saved successfully: {excel_filename}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
    
    print("\nAnalysis complete!")

if __name__ == "__main__":
    main()